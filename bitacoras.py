import streamlit as st
import pandas as pd
from datetime import datetime, date
import io
from base_datos import conectar

def mostrar(supervisor_id=None):
    # Estilo CSS limpio para matrices sin desperdiciar espacio vertical
    st.markdown("""
        <style>
        .block-header { background-color: #E5E7EB; padding: 6px; font-weight: bold; color: #1F2937; border-left: 6px solid #4B5563; margin-top: 1rem; margin-bottom: 0.5rem; font-size: 14px; }
        .stDataEditor { font-size: 11px; }
        div[data-testid="stExpander"] div { padding-top: 0px; }
        </style>
    """, unsafe_allow_html=True)

    supabase = conectar()
    
    # Manejo de estados persistentes de la sesión
    if 'id_bitacora_activa' not in st.session_state:
        st.session_state.id_bitacora_activa = None

    # Si hay una bitácora abierta para edición, mostramos la interfaz unificada de 4 secciones
    if st.session_state.id_bitacora_activa:
        renderizar_formato_edicion(st.session_state.id_bitacora_activa, supabase)
    else:
        # Menú de pestañas principales optimizado sin títulos redundantes
        tab_historial, tab_alta = st.tabs(["🗂️ Listado de Bitácoras", "➕ Nueva Bitácora"])

        # =========================================================================
        # PESTAÑA A: HISTORIAL, BÚSQUEDA Y CAMBIO DE ESTADOS
        # =========================================================================
        with tab_historial:
            busq = st.text_input("🔍 Filtrar historial:", placeholder="Escriba Nº Orden, Proyecto, Cliente o Mueble para buscar...")
            
            try:
                res_t = supabase.table("bitacoras_taller").select("*").execute()
                df_t = pd.DataFrame(res_t.data) if res_t.data else pd.DataFrame()
                if not df_t.empty and 'fecha' in df_t.columns:
                    df_t = df_t.sort_values(by="fecha", ascending=False)
            except Exception as e:
                st.error(f"Error de comunicación con Supabase: {e}")
                df_t = pd.DataFrame()

            if busq and not df_t.empty:
                df_t = df_t[
                    df_t['n_orden'].astype(str).str.contains(busq, case=False) |
                    df_t['proyecto'].astype(str).str.contains(busq, case=False) |
                    df_t['cliente'].astype(str).str.contains(busq, case=False) |
                    df_t['tipo_mueble'].astype(str).str.contains(busq, case=False)
                ]

            if not df_t.empty:
                # Modificación rápida de estado directo sobre la grilla interactiva
                st.caption("💡 Puede cambiar el Estado de avance directamente en la tabla y presionar Guardar Estado.")
                
                df_editor_estados = st.data_editor(
                    df_t[['id', 'fecha', 'n_orden', 'proyecto', 'cliente', 'tipo_mueble', 'estado']],
                    column_config={
                        "id": st.column_config.TextColumn("ID", disabled=True),
                        "fecha": st.column_config.TextColumn("Fecha", disabled=True),
                        "n_orden": st.column_config.TextColumn("Nº Orden", disabled=True),
                        "proyecto": st.column_config.TextColumn("Proyecto", disabled=True),
                        "cliente": st.column_config.TextColumn("Cliente", disabled=True),
                        "tipo_mueble": st.column_config.TextColumn("Tipo Mueble", disabled=True),
                        "estado": st.column_config.SelectboxColumn("Estado Actual", options=["Pendiente", "En Proceso", "Cerrada"], required=True)
                    },
                    hide_index=True,
                    use_container_width=True,
                    key="editor_estados_directo"
                )
                
                c_est1, c_est2 = st.columns([1, 2])
                if c_est1.button("💾 Guardar Estados Modificados", use_container_width=True):
                    try:
                        for idx, r_est in df_editor_estados.iterrows():
                            supabase.table("bitacoras_taller").update({"estado": r_est['estado']}).eq("id", int(r_est['id'])).execute()
                        st.success("Estados actualizados correctamente.")
                        st.rerun()
                    except Exception as e:
                        st.error(f"Falla al actualizar lotes: {e}")

                st.divider()
                
                # Apertura controlada por ID
                id_sel = st.number_input("Indique el ID de la Bitácora que desea reabrir para actualización:", min_value=1, step=1, key="num_id_abrir")
                if st.button("🔓 Abrir Formato en Pantalla", type="primary", use_container_width=True):
                    if id_sel in df_t['id'].tolist():
                        st.session_state.id_bitacora_activa = int(id_sel)
                        st.rerun()
                    else:
                        st.error("El ID indicado no corresponde a ningún documento.")
            else:
                st.info("No se registran documentos de bitácora bajo los parámetros ingresados.")

        # =========================================================================
        # PESTAÑA B: ALTA NUEVA (CON CAMPOS PARCIALES O VACÍOS)
        # =========================================================================
        with tab_alta:
            st.caption("Habilite una nueva bitácora para la jornada. Puede rellenar únicamente la primera sección y dejar el resto vacío.")
            with st.form("alta_bitacora_modular_form"):
                c1, c2 = st.columns(2)
                f_doc = c1.date_input("FECHA:", value=date.today())
                n_ord = c2.text_input("Nº ORDEN:")
                
                c3, c4 = st.columns(2)
                t_mue = c3.text_input("TIPO DE MUEBLE:")
                mot = c4.text_input("MOTIVO:")
                
                c5, c6 = st.columns(2)
                cli = c5.text_input("CLIENTE:")
                proy = c6.text_input("PROYECTO:")
                
                c7, c8 = st.columns(2)
                sol_por = c7.text_input("SOLICITADO POR:")
                sup_prod = c8.text_input("SUP. DE PRODUCCION:", value="DOMÉNICO MORÓN")
                
                if st.form_submit_button("🚀 Inicializar Hoja (Por Defecto: Pendiente)", type="primary", use_container_width=True):
                    res_ins = supabase.table("bitacoras_taller").insert({
                        "fecha": f_doc.isoformat(), "n_orden": n_ord, "tipo_mueble": t_mue,
                        "motivo": mot, "cliente": cli, "proyecto": proy,
                        "solicitado_por": sol_por, "sup_production": sup_prod, "estado": "Pendiente"
                    }).execute()
                    
                    if res_ins.data:
                        b_id = res_ins.data[0]['id']
                        
                        # Pre-cargar filas en blanco reglamentarias por proceso (3 por máquina)
                        lineas_iniciales = []
                        for _ in range(3):
                            lineas_iniciales.append({"bitacora_id": b_id, "proceso_bloque": "SECCIONADORA", "cantidad": 0.0})
                            lineas_iniciales.append({"bitacora_id": b_id, "proceso_bloque": "ESCUADRADORA", "cantidad": 0.0})
                            lineas_iniciales.append({"bitacora_id": b_id, "proceso_bloque": "CANTEO", "cantidad": 0.0})
                            
                        supabase.table("bitacoras_lineas").insert(lineas_iniciales).execute()
                        st.session_state.id_bitacora_activa = b_id
                        st.rerun()


def renderizar_formato_edicion(id_act, supabase):
    """Módulo interno que dibuja exactamente las 4 secciones en pantalla de forma simétrica"""
    if st.button("⬅️ Cerrar Edición y Volver al Historial"):
        st.session_state.id_bitacora_activa = None
        st.rerun()

    # Cargar cabecera
    cab = supabase.table("bitacoras_taller").select("*").eq("id", id_act).execute().data[0]
    
    # SECCIÓN 1: DATOS GENERALES (Imagen 1)
    st.markdown('<div class="block-header">📄 SECCIÓN 1: DATOS GENERALES DEL FORMATO</div>', unsafe_allow_html=True)
    with st.container(border=True):
        col1, col2 = st.columns(2)
        u_fecha = col1.date_input("FECHA:", value=datetime.strptime(cab['fecha'], "%Y-%m-%d").date())
        u_n_orden = col2.text_input("Nº ORDEN:", value=cab['n_orden'] or "")
        
        u_tipo_mueble = col1.text_input("TIPO DE MUEBLE:", value=cab['tipo_mueble'] or "")
        u_motivo = col2.text_input("MOTIVO:", value=cab['motivo'] or "")
        
        u_cliente = col1.text_input("CLIENTE:", value=cab['cliente'] or "")
        u_proyecto = col2.text_input("PROYECTO:", value=cab['proyecto'] or "")
        
        u_sol_por = col1.text_input("SOLICITADO POR:", value=cab['solicitado_por'] or "")
        u_sup_prod = col2.text_input("SUP. DE PRODUCCION:", value=cab['sup_production'] or "")
        
        u_estado = st.selectbox("Cambiar Estado de la Hoja Activa:", ["Pendiente", "En Proceso", "Cerrada"], index=["Pendiente", "En Proceso", "Cerrada"].index(cab['estado']))

    # Cargar las líneas de detalle de la base de datos
    res_l = supabase.table("bitacoras_lineas").select("*").eq("bitacora_id", id_act).order("id").execute()
    df_l = pd.DataFrame(res_l.data) if res_l.data else pd.DataFrame()
    
    df_secc = df_l[df_l['proceso_bloque'] == 'SECCIONADORA'].copy() if not df_l.empty else pd.DataFrame()
    df_escu = df_l[df_l['proceso_bloque'] == 'ESCUADRADORA'].copy() if not df_l.empty else pd.DataFrame()
    df_cant = df_l[df_l['proceso_bloque'] == 'CANTEO'].copy() if not df_l.empty else pd.DataFrame()

    # SECCIÓN 2: CORTE SECCIONADORA (Imagen 2)
    st.markdown('<div class="block-header">🪚 SECCIÓN 2: CORTE SECCIONADORA</div>', unsafe_allow_html=True)
    if st.button("➕ Insertar Registro a Seccionadora"):
        supabase.table("bitacoras_lineas").insert({"bitacora_id": id_act, "proceso_bloque": "SECCIONADORA", "cantidad": 0.0}).execute()
        st.rerun()
        
    ed_secc = st.data_editor(
        df_secc[['id', 'cantidad', 'descripcion', 'fecha_inicio', 'hora_inicio', 'cant_final_pl_pzs', 'hora_termino', 'fecha_termino', 'obs_incidencias', 'nombre_firma_operario']],
        column_config={
            "id": None,
            "cantidad": st.column_config.NumberColumn("CANT.", format="%.2f"),
            "descripcion": st.column_config.TextColumn("DESCRIPCION"),
            "fecha_inicio": st.column_config.TextColumn("FECHA INICIO"),
            "hora_inicio": st.column_config.TextColumn("HORA INICIO"),
            "cant_final_pl_pzs": st.column_config.TextColumn("CANT. FINAL PL."),
            "hora_termino": st.column_config.TextColumn("HORA TERMINO"),
            "fecha_termino": st.column_config.TextColumn("FECHA TERMINO"),
            "obs_incidencias": st.column_config.TextColumn("OBS/INCIDENCIAS"),
            "nombre_firma_operario": st.column_config.TextColumn("NOMBRE Y FIRMA DE OPERARIO CORTE")
        },
        hide_index=True, use_container_width=True, key=f"matriz_secc_{id_act}"
    )

    # SECCIÓN 3: CORTE ESCUADRADORA (Imagen 3)
    st.markdown('<div class="block-header">📐 SECCIÓN 3: CORTE ESCUADRADORA</div>', unsafe_allow_html=True)
    if st.button("➕ Insertar Registro a Escuadradora"):
        supabase.table("bitacoras_lineas").insert({"bitacora_id": id_act, "proceso_bloque": "ESCUADRADORA", "cantidad": 0.0}).execute()
        st.rerun()
        
    ed_escu = st.data_editor(
        df_escu[['id', 'cantidad', 'descripcion', 'fecha_inicio', 'hora_inicio', 'cant_final_pl_pzs', 'hora_termino', 'fecha_termino', 'obs_incidencias', 'nombre_firma_operario']],
        column_config={
            "id": None,
            "cantidad": st.column_config.NumberColumn("CANT.", format="%.2f"),
            "descripcion": st.column_config.TextColumn("DESCRIPCION"),
            "fecha_inicio": st.column_config.TextColumn("FECHA INICIO"),
            "hora_inicio": st.column_config.TextColumn("HORA INICIO"),
            "cant_final_pl_pzs": st.column_config.TextColumn("CANT. PIEZAS"),
            "hora_termino": st.column_config.TextColumn("HORA TERMINO"),
            "fecha_termino": st.column_config.TextColumn("FECHA TERMINO"),
            "obs_incidencias": st.column_config.TextColumn("OBS/INCIDENCIAS"),
            "nombre_firma_operario": st.column_config.TextColumn("NOMBRE Y FIRMA DE OPERARIO CORTE")
        },
        hide_index=True, use_container_width=True, key=f"matriz_escu_{id_act}"
    )

    # SECCIÓN 4: CANTEO (Imagen 4)
    st.markdown('<div class="block-header">⚙️ SECCIÓN 4: CANTEO</div>', unsafe_allow_html=True)
    if st.button("➕ Insertar Registro a Canteo"):
        supabase.table("bitacoras_lineas").insert({"bitacora_id": id_act, "proceso_bloque": "CANTEO", "cantidad": 0.0}).execute()
        st.rerun()
        
    ed_cant = st.data_editor(
        df_cant[['id', 'cantidad', 'descripcion', 'tipo_canto', 'fecha_inicio', 'hora_inicio', 'cant_final_pl_pzs', 'fecha_termino', 'obs_incidencias', 'nombre_firma_operario']],
        column_config={
            "id": None,
            "cantidad": st.column_config.NumberColumn("CANT.", format="%.2f"),
            "descripcion": st.column_config.TextColumn("DESCRIPCION"),
            "tipo_canto": st.column_config.TextColumn("TIPO DE CANTO"),
            "fecha_inicio": st.column_config.TextColumn("FECHA INICIO"),
            "hora_inicio": st.column_config.TextColumn("HORA INICIAL"),
            "cant_final_pl_pzs": st.column_config.TextColumn("CANTO USADO"),
            "fecha_termino": st.column_config.TextColumn("FECHA FINAL"),
            "obs_incidencias": st.column_config.TextColumn("OBS/INCIDENCIAS"),
            "nombre_firma_operario": st.column_config.TextColumn("NOMBRE Y FIRMA DE OPERARIO CORTE")
        },
        hide_index=True, use_container_width=True, key=f"matriz_cant_{id_act}"
    )

    # BARRA CENTRAL DE CONTROLES FINALES
    st.divider()
    b_col1, b_col2 = st.columns(2)
    
    if b_col1.button("💾 GUARDAR CAMBIOS Y AVANCES EN LA APP", type="primary", use_container_width=True):
        try:
            # 1. Salvar cambios de la Sección 1 (Cabecera)
            supabase.table("bitacoras_taller").update({
                "fecha": u_fecha.isoformat(), "n_orden": u_n_orden, "tipo_mueble": u_tipo_mueble,
                "motivo": u_motivo, "cliente": u_cliente, "proyecto": u_proyecto,
                "solicitado_por": u_sol_por, "sup_production": u_sup_prod, "estado": u_estado
            }).eq("id", id_act).execute()
            
            # 2. Consolidar matrices horizontales
            todo_detalle = (
                ed_secc.to_dict(orient='records') + 
                ed_escu.to_dict(orient='records') + 
                ed_cant.to_dict(orient='records')
            )
            
            for fila in todo_detalle:
                supabase.table("bitacoras_lineas").update({
                    "cantidad": float(fila['cantidad']) if fila['cantidad'] else 0.0,
                    "descripcion": str(fila['descripcion']).strip() if fila['descripcion'] else None,
                    "tipo_canto": str(fila.get('tipo_canto', '')).strip() if fila.get('tipo_canto') else None,
                    "fecha_inicio": str(fila['fecha_inicio']).strip() if fila['fecha_inicio'] else None,
                    "hora_inicio": str(fila['hora_inicio']).strip() if fila['hora_inicio'] else None,
                    "cant_final_pl_pzs": str(fila['cant_final_pl_pzs']).strip() if fila['cant_final_pl_pzs'] else None,
                    "hora_termino": str(fila.get('hora_termino', '')).strip() if fila.get('hora_termino') else None,
                    "fecha_termino": str(fila['fecha_termino']).strip() if fila['fecha_termino'] else None,
                    "obs_incidencias": str(fila['obs_incidencias']).strip() if fila['obs_incidencias'] else None,
                    "nombre_firma_operario": str(fila['nombre_firma_operario']).strip() if fila['nombre_firma_operario'] else None
                }).eq("id", int(fila['id'])).execute()
                
            st.success("🎉 Bitácora guardada y actualizada con éxito."); st.cache_data.clear(); st.rerun()
        except Exception as e:
            st.error(f"Falla de guardado en Supabase: {e}")

    # EXPORTACIÓN ESTRUCTURADA A EXCEL CON DISEÑO INDUSTRIAL (openpyxl)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TRAZABILIDAD"
    ws.views.sheetView[0].showGridLines = True
    
    # Definición de Estilos Oficiales
    f_cab = Font(name="Arial", size=10, bold=True, color="000000")
    f_tit = Font(name="Arial", size=14, bold=True, color="000000")
    f_reg = Font(name="Arial", size=9, bold=False)
    fill_gris = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center")
    border_fino = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # CONSTRUCCIÓN DE LA SECCIÓN 1 (CABECERA GRIS)
    datos_s1 = [
        ("FECHA:", cab['fecha'], "Nº ORDEN:", cab['n_orden']),
        ("TIPO DE MUEBLE:", cab['tipo_mueble'], "MOTIVO:", cab['motivo']),
        ("CLIENTE:", cab['cliente'], "PROYECTO:", cab['proyecto']),
        ("SOLICITADO POR:", cab['solicitado_por'], "SUP. DE PRODUCCION:", cab['sup_production'])
    ]
    
    for r_idx, row_data in enumerate(datos_s1, start=1):
        ws.cell(row=r_idx, column=1, value=row_data[0]).font = f_cab
        ws.cell(row=r_idx, column=1).fill = fill_gris
        ws.cell(row=r_idx, column=2, value=row_data[1]).font = f_reg
        ws.cell(row=r_idx, column=3, value=row_data[2]).font = f_cab
        ws.cell(row=r_idx, column=3).fill = fill_gris
        ws.cell(row=r_idx, column=4, value=row_data[3]).font = f_reg
        for c in range(1, 5):
            ws.cell(row=r_idx, column=c).border = border_fino

    cursor = 6

    # FUNCIÓN INTERNA PARA AGREGAR TABLAS OPERATIVAS
    def inyectar_bloque_excel(ws, start_row, titulo, cabeceras, df_datos, es_canteo=False):
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(cabeceras))
        cell_t = ws.cell(row=start_row, column=1, value=titulo)
        cell_t.font = f_tit
        cell_t.alignment = align_center
        cell_t.fill = fill_gris
        ws.row_dimensions[start_row].height = 25
        
        # Pintar cabeceras
        header_row = start_row + 1
        for c_idx, text_h in enumerate(cabeceras, start=1):
            cell_h = ws.cell(row=header_row, column=c_idx, value=text_h)
            cell_h.font = f_cab
            cell_h.fill = fill_gris
            cell_h.alignment = align_center
            cell_h.border = border_fino
        ws.row_dimensions[header_row].height = 20
        
        idx_act = header_row + 1
        if not df_datos.empty:
            for _, r in df_datos.iterrows():
                ws.row_dimensions[idx_act].height = 18
                if not es_canteo:
                    vals = [r['cantidad'], r['descripcion'], r['fecha_inicio'], r['hora_inicio'], r['cant_final_pl_pzs'], r['hora_termino'], r['fecha_termino'], r['obs_incidencias'], r['nombre_firma_operario']]
                else:
                    vals = [r['cantidad'], r['descripcion'], r['tipo_canto'], r['fecha_inicio'], r['hora_inicio'], r['cant_final_pl_pzs'], r['fecha_termino'], r['obs_incidencias'], r['nombre_firma_operario']]
                
                for col_i, val in enumerate(vals, start=1):
                    cell_d = ws.cell(row=idx_act, column=col_i, value=val if val is not None else "")
                    cell_d.font = f_reg
                    cell_d.border = border_fino
                    if col_i in [1, 3, 4, 5, 6, 7]: cell_d.alignment = align_center
                idx_act += 1
        else:
            # Si está vacío el registro, inyectamos 3 filas limpias reglamentarias para impresión manual
            for _ in range(3):
                ws.row_dimensions[idx_act].height = 18
                for col_i in range(1, len(cabeceras) + 1):
                    ws.cell(row=idx_act, column=col_i).border = border_fino
                idx_act += 1
                
        # Agregar el pie de firma del supervisor de producción al final de cada máquina
        ws.cell(row=idx_act, column=len(cabeceras)-1, value="VºBº SUP. PRODUCCION").font = f_cab
        ws.cell(row=idx_act, column=len(cabeceras)-1).fill = fill_gris
        ws.cell(row=idx_act, column=len(cabeceras)-1).border = border_fino
        ws.cell(row=idx_act, column=len(cabeceras)).border = border_fino
        ws.row_dimensions[idx_act].height = 20
        
        return idx_act + 2

    # Inyección consecutiva de los tres bloques mecánicos
    cursor = inyectar_bloque_excel(ws, cursor, "CORTE SECCIONADORA", ["CANT.", "DESCRIPCION", "FECHA INICIO", "HORA INICIO", "CANT. FINAL PL.", "HORA TERMINO", "FECHA TERMINO", "OBS/INCIDENCIAS", "NOMBRE Y FIRMA DE OPERARIO CORTE"], ed_secc)
    cursor = inyectar_bloque_excel(ws, cursor, "CORTE ESCUADRADORA", ["CANT.", "DESCRIPCION", "FECHA INICIO", "HORA INICIO", "CANT. PIEZAS", "HORA TERMINO", "FECHA TERMINO", "OBS/INCIDENCIAS", "NOMBRE Y FIRMA DE OPERARIO CORTE"], ed_escu)
    cursor = inyectar_bloque_excel(ws, cursor, "CANTEO", ["CANT.", "DESCRIPCION", "TIPO DE CANTO", "FECHA INICIO", "HORA INICIAL", "CANTO USADO", "FECHA FINAL", "OBS/INCIDENCIAS", "NOMBRE Y FIRMA DE OPERARIO CORTE"], ed_cant, es_canteo=True)

    # Auto-ajustar anchos de columnas del Excel para que no se corten los textos al mandar a la impresora
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    out_excel = io.BytesIO()
    wb.save(out_excel)
    
    b_col2.download_button(
        "🖨️ EXPORTAR BITÁCORA FIEL (EXCEL/IMPRESIÓN)",
        data=out_excel.getvalue(),
        file_name=f"Bitacora_Trazabilidad_{cab['id']}_{cab['n_orden']}.xlsx",
        use_container_width=True
    )
